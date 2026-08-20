#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importa Equipos_2026.xlsx y Ciclistas_2026_1.xlsx a data/seed.db (SQLite).

Uso:
    python tools/import_data.py

Notas:
- Vincula corredores con equipos por NOMBRE normalizado (los TeamID de ambos
  archivos no coinciden).
- Genera colores deterministas (principal/secundario) por equipo.
- El esquema resultante es la fuente de la verdad para la app Godot.
"""

import hashlib
import os
import re
import sqlite3
import unicodedata
from colorsys import hls_to_rgb

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EQUIPOS = os.path.join(BASE, "Equipos_2026.xlsx")
CICLISTAS = os.path.join(BASE, "Ciclistas_2026_1.xlsx")
OUT_DB = os.path.join(BASE, "data", "seed.db")

ATTRS = ["fla", "mnt", "mm", "hil", "ttr", "prl", "cob", "spr",
         "acc", "dhi", "att", "sta", "res", "rec"]


def normalize(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.casefold()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def color_for(name):
    digest = hashlib.md5(name.encode("utf-8")).digest()
    hue = (digest[0] * 256 + digest[1]) % 360
    prgb = hls_to_rgb(hue / 360.0, 0.55, 0.62)
    srgb = hls_to_rgb(((hue + 150) % 360) / 360.0, 0.30, 0.55)

    def hexc(rgb):
        return "#" + "".join("%02X" % int(c * 255) for c in rgb)

    return hexc(prgb), hexc(srgb)


SCHEMA = """
CREATE TABLE IF NOT EXISTS teams (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    abbr TEXT,
    country TEXT,
    category TEXT,
    color_primary TEXT,
    color_secondary TEXT,
    extra INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS riders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    birth_date TEXT,
    nationality TEXT,
    team_id INTEGER REFERENCES teams(id),
    specialty TEXT,
    fla INTEGER, mnt INTEGER, mm INTEGER, hil INTEGER,
    ttr INTEGER, prl INTEGER, cob INTEGER, spr INTEGER,
    acc INTEGER, dhi INTEGER, att INTEGER, sta INTEGER,
    res INTEGER, rec INTEGER
);
CREATE TABLE IF NOT EXISTS stages (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    date TEXT,
    type TEXT,
    distance REAL,
    start TEXT,
    finish TEXT,
    description TEXT,
    sections_json TEXT,
    modifiers_json TEXT,
    difficulty INTEGER DEFAULT 0,
    locked INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS races (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    edition TEXT,
    country TEXT,
    description TEXT,
    start_date TEXT,
    end_date TEXT,
    logo TEXT,
    stage_order_json TEXT
);
CREATE TABLE IF NOT EXISTS simulations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    date TEXT,
    seed TEXT,
    mode TEXT,
    ref_type TEXT,
    ref_id INTEGER,
    results_json TEXT,
    classifications_json TEXT,
    events_json TEXT,
    decisions_json TEXT
);
CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT
);
"""


def load_teams():
    import openpyxl
    wb = openpyxl.load_workbook(EQUIPOS, data_only=True)
    ws = wb.active
    teams = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        if row[2] is None:
            continue
        name = str(row[2]).strip()
        teams.append({
            "name": name,
            "abbr": str(row[3]).strip() if row[3] else None,
            "country": str(row[4]).strip() if row[4] else None,
            "category": str(row[5]).strip() if row[5] else None,
            "extra": 0,
        })
    return teams


def load_riders():
    import openpyxl
    wb = openpyxl.load_workbook(CICLISTAS, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    riders = []
    for r in rows:
        if r[0] is None:
            continue
        attr_values = []
        for i in range(7, 7 + len(ATTRS)):
            v = r[i]
            attr_values.append(int(v) if v is not None else None)
        riders.append({
            "name": str(r[0]).strip(),
            "birth_date": r[1],
            "nationality": str(r[2]).strip() if r[2] else None,
            "team_name": str(r[3]).strip() if r[3] else None,
            "specialty": r[6],
            "attrs": dict(zip(ATTRS, attr_values)),
        })
    return riders


def main():
    teams = load_teams()
    riders = load_riders()

    os.makedirs(os.path.dirname(OUT_DB), exist_ok=True)
    if os.path.exists(OUT_DB):
        os.remove(OUT_DB)
    conn = sqlite3.connect(OUT_DB)
    cur = conn.cursor()
    cur.executescript(SCHEMA)

    # Insertar equipos (dedupe por nombre normalizado).
    team_id_by_norm = {}
    seen = set()
    extra_teams = {}
    for t in teams:
        norm = normalize(t["name"])
        if norm in seen:
            continue
        seen.add(norm)
        cp, cs = color_for(t["name"])
        cur.execute(
            "INSERT INTO teams (name, abbr, country, category, color_primary, color_secondary, extra)"
            " VALUES (?,?,?,?,?,?,?)",
            (t["name"], t["abbr"], t["country"], t["category"], cp, cs, 0),
        )
        team_id_by_norm[norm] = cur.lastrowid

    # Equipos referenciados por corredores pero ausentes en Equipos.xlsx.
    for r in riders:
        norm = normalize(r["team_name"])
        if norm and norm not in team_id_by_norm:
            cp, cs = color_for(r["team_name"])
            cur.execute(
                "INSERT INTO teams (name, abbr, country, category, color_primary, color_secondary, extra)"
                " VALUES (?,?,?,?,?,?,?)",
                (r["team_name"], None, None, "Continental", cp, cs, 1),
            )
            team_id_by_norm[norm] = cur.lastrowid

    # Corredores.
    for r in riders:
        norm = normalize(r["team_name"])
        team_id = team_id_by_norm.get(norm)
        cur.execute(
            "INSERT INTO riders (name, birth_date, nationality, team_id, specialty,"
            " fla, mnt, mm, hil, ttr, prl, cob, spr, acc, dhi, att, sta, res, rec)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                r["name"], r["birth_date"], r["nationality"], team_id, r["specialty"],
                r["attrs"]["fla"], r["attrs"]["mnt"], r["attrs"]["mm"], r["attrs"]["hil"],
                r["attrs"]["ttr"], r["attrs"]["prl"], r["attrs"]["cob"], r["attrs"]["spr"],
                r["attrs"]["acc"], r["attrs"]["dhi"], r["attrs"]["att"], r["attrs"]["sta"],
                r["attrs"]["res"], r["attrs"]["rec"],
            ),
        )

    conn.commit()

    n_teams = cur.execute("SELECT COUNT(*) FROM teams").fetchone()[0]
    n_extra = cur.execute("SELECT COUNT(*) FROM teams WHERE extra=1").fetchone()[0]
    n_riders = cur.execute("SELECT COUNT(*) FROM riders").fetchone()[0]
    n_unlinked = cur.execute("SELECT COUNT(*) FROM riders WHERE team_id IS NULL").fetchone()[0]
    conn.close()

    print("Importación completada -> %s" % OUT_DB)
    print("  Equipos: %d (%d extra, sin metadatos)" % (n_teams, n_extra))
    print("  Corredores: %d" % n_riders)
    print("  Corredores sin equipo: %d" % n_unlinked)


if __name__ == "__main__":
    main()
